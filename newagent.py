from openpyxl import load_workbook

print("I am your Excel working agent!")

filename = input("Enter your file name (e.g., my_data.xlsx): ")

try:
    workbook = load_workbook(filename)
    page = workbook.active
    
    print("\nI am ready, Type 'done' to save and exit.")
    
    while True: #This mans it will run until the user types 'done'
        user_action = input("What kind of data settlement do you want? (double/increase/add/sort/search/pricing/done): ").lower()

        if user_action == "done":
            print("\nSaving changes and exiting.")
            break 

        elif user_action == "double":
            print("Processing double values...")
            for row_index, row in enumerate(page.iter_rows(min_row=2), start=2):
                pieces_left_cell = row[1]#ye column ki form mein h jese pieces, product and barcode u yeh row 1 h!!!
                if isinstance(pieces_left_cell.value, (int, float)):#isinstance check krta h ky value integar ya float ya koi text!
                    if pieces_left_cell.value < 300:
                        pieces_left_cell.value *= 2
                        print(f"Row {row_index} doubled to {pieces_left_cell.value}.")
                    else:
                        print(f"Row {row_index}:")
                else:
                    print(f"Row {row_index}: Skip non-numeric value '{pieces_left_cell.value}'.")
        
        elif user_action == "increase":
            item = input("Enter the item you wanna increase").strip().lower()
            quantity = int(input("Enter the quantity how much you wanna increase!!"))
            print(f"Processing data to increase '{item.title()}' by {quantity}..")

            for row_index, row in enumerate(page.iter_rows(min_row=2), start=2):
                item_name = str(row[0].value).strip().lower() 
                pieces_left_cell = row[1]
              
                if item_name == item:
                    if isinstance(pieces_left_cell.value, (int, float)):
                      pieces_left_cell.value += quantity
                      print(f"Row {row_index} - Item '{item_name.title()}' increased by {quantity}. New value: {pieces_left_cell.value}")
                    else:
                      print(f"\n '{row_index}' skipping non-numeric values '{pieces_left_cell}'.")

        elif user_action == "add":
            print("Processing to add new data...")
            adding_item = input("Enter the item to add")
            quantity_add = int(input("Enter the quantity.."))
            barcode_add = int(input("Enter the barcode"))
            price_add = int(input("Enter the price"))
            new_data = [adding_item, quantity_add, barcode_add, price_add]
            page.append(new_data)
            print(f"The product is added successfully in inventory '{new_data}'.")

        
        elif user_action == "sort":
              print(f"Processing to add new data..")
              rows = list(page.iter_rows(min_row=2, values_only=True))
              rows.sort(key=lambda x :x[1], reverse=True)#descending order

              for row in page.iter_rows(min_row=2):#for deletion of rows 
                  for cell in row:
                      cell.value = None
                      
              for row_data in rows:#iska mtlb h ky rows joky sort hochuki hn usmy se row_data rows mein sy piece piece krky uthaty hn
                  page.append(list(row_data))

              print("Data sorted successfully!")

        elif user_action == "search":
         print("\nSearching is processing....")
         product_name = input("Enter the product name to search: ").lower()

         for row in page.iter_rows(min_row=2):
        
           product_cell = row[0]
           pieces_cell = row[1]
           barcode_cell = row[2]
           price_cell = row[3]

           if product_cell.value is not None and str(product_cell.value).lower() == product_name:
            
             print("\n--- Product Found! ---")
             print(f"Product: {product_cell.value}")
             print(f"Pieces Left: {pieces_cell.value}")
             print(f"Barcode: {barcode_cell.value}")
             print(f"Price: {price_cell.value}")
             break 
         else: 
                print(f"\nSorry, the product '{product_name}' was not found in the sheet.")
           
        elif user_action == "price":
         product_name = input("Enter the product name to find the price for: ").lower()
    
         found = False
        
         for row in page.iter_rows(min_row=2):
        
           product_cell = row[0]
           price_cell = row[3]
        
           if product_cell.value is not None and str(product_cell.value).lower() == product_name:
            
             if isinstance(price_cell.value, (int, float)):
                
                print("\n--- Price Found! ---")
                print(f"The price for {product_cell.value} is PKR {price_cell.value}")
                
                found = True
                break
             else:
                print(f"\nError: Price found for {product_cell.value}, but the value is not a valid number.")
                found = True
                break

         if not found:
           print(f"\nSorry, the product '{product_name}' was not found in the sheet.")

        elif user_action == "lowstockproducts":
           print("\n--- Generating Low Stock Report ---")

           stock_limit = int(input("Enter the low stock limit (e.g., 50): "))
           stock_limit = 100

        low_stock_items = []
        for row in page.iter_rows(min_row=2):
           product = row[0].value
           pieces = row[1].value

            
        if isinstance(pieces, (int, float)) and pieces is not None and pieces < stock_limit:
           low_stock_items.append({
                'product': product,
                'pieces': pieces
            })
           
           print(f"\nReport for Items Below {stock_limit} units:")
           if low_stock_items:
              for item in low_stock_items:
                  print(f"LOW STOCK: {item['product']} (Only {item['pieces']} left!)")
              else:
                  print("Great! All products are above the specified stock limit.")
               
    new_filename = 'updated_' + filename
    workbook.save(new_filename)
    print(f"\nDone! The updated data has been saved to '{new_filename}'.")

except FileNotFoundError:
    print(f"\nError: The file '{filename}' was not found. Make sure it's in the same folder.")
except Exception as e:
    print(f"\nAn unexpected error occurred: {e}")

finally:
    print(f"Succeed!!!")